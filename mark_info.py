import requests
import time
import random
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
url = 'https://mops.twse.com.tw/mops/web/ajax_t100sb14'

url1 = 'https://mops.twse.com.tw/mops/web/t100sb14'

#尋找公司代號
def find_index(companys,as_companyno):      
    li_return = 0
    i = 0
    for company in companys:
        if company[1] == as_companyno:
           li_return = i
           break
        i = i + 1
    return li_return


#取得市場別id，其實就是上市及上櫃的值
def get_market():
    market_to_id = {}
    resp = requests.get(url1)
    soup = BeautifulSoup(resp.text, 'html5lib')
    selects = soup.find_all('select',{'name':'code'})
    options = soup.find('select',{'name':'TYPEK'}).find_all('option')
    i = 0
    for opt in options:
        market_to_id[opt.text.strip()] = opt['value']
    return market_to_id 

def gen_companylist():
    #==讀取上市及上櫃所有資料==========================================
    companys = list()
    title = list()
    wb1 = load_workbook('上市公司資料.xlsx')
    wb2 = load_workbook('上櫃公司資料.xlsx')
    ws1 = wb1.active
    ws2 = wb2.active
    
    i = 0
    for row in ws1.rows:
        company = list()
        if i == 0:
            for col in row:
                title.append(col.value)
            title.append('員工福利費用(仟元)')
            title.append('員工薪資費用(仟元)')            
            title.append('員工人數')
            title.append('平均員工福利費用(仟元/人)')  
            title.append('平均員工薪資費用(仟元/人)')
            title.append('每股盈餘(元/股)')  
            title.append('同產業平均員工福利費用(仟元/人)')
            title.append('同產業平均員工薪資費用(仟元/人)')              
            title.append('同產業平均每股盈餘(元/股)')     
            companys.append(title)  #加入上市公司資料    
        else:
            for col in row:
                company.append(col.value)
            companys.append(company)  #加入上市公司資料
        i = i + 1      
        

    i = 0
    for row in ws2.rows:
        company = list()
        if i == 0:
            #Do Nothing
            i == i
        else:
            for col in row:
                company.append(col.value)
            companys.append(company)  #加入上櫃公司資料
        i = i + 1 
    #===============================================================

    wb1.close()    #關閉工作簿
    wb2.close()    #關閉工作簿   
    return companys


#===========================主程式==================================
if __name__ == '__main__':   
    companys = gen_companylist() #抓取公司List
    #抓取網路相關資料
    market = get_market() #取得上市或是上櫃
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36'}
    
    for x in market:
        # 只抓股票名稱那欄  如果不給code數值，就是全部公司抓取，maket[x]代表上市或是上櫃公司
        d = {
            'encodeURIComponent':'1',
            'step':'1',
            'firstin':'1',
            'TYPEK': market[x],
            'RYEAR':'107',
            'code':''
            }
        r = requests.post(url, data=d,headers = headers)
        soup = BeautifulSoup(r.text,'html.parser')
        
        table = soup.find('table')
        rows = table.find_all('tr')
        #====Merger Two Lists==================================================
        i = 0
        for row in rows:
            if i <= 1:
                i = i
            else:
                company = list()
                for column in row.find_all('td'):
                    company.append(column.text.strip().replace(',',''))
                    #搜尋companys
                li_id = find_index(companys,int(company[1]))
                j = 0
                for col in company:
                    if j >= 4: #從第五個欄位開始抓取
                       companys[li_id].append(col)
                    j = j + 1
                                
            i = i + 1

#======寫入Excel檔案======================
        wb3 = Workbook() #產生新的EXCEL
        ws3 = wb3.active
        for com in companys:
            ws3.append(list(com))           
        wb3.save('上市上櫃總名單.xlsx')                
        wb3.close()
        print('寫入資料' + market[x])
        print('休息10秒...')
        time.sleep(10)
        #休息10秒繼續抓，避免被網站認為是機器人攻擊
